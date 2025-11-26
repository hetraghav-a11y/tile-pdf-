import os
import uuid
from io import BytesIO
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect,
    send_file, flash
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from PIL import Image
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader

# ------------------------------------------------
# App + DB setup
# ------------------------------------------------
app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = "dev_secret_for_flashes"

db = SQLAlchemy(app)

# Folders (all under project root)
os.makedirs(os.path.join(app.root_path, "static", "images"), exist_ok=True)
os.makedirs(os.path.join(app.root_path, "static", "logo"), exist_ok=True)
os.makedirs(os.path.join(app.root_path, "static", "posters"), exist_ok=True)
os.makedirs(os.path.join(app.root_path, "static", "tile_templates"), exist_ok=True)
os.makedirs(os.path.join(app.root_path, "uploads"), exist_ok=True)

POSTER_DIR = os.path.join(app.root_path, "static", "posters")
TILE_TEMPLATE_DIR = os.path.join(app.root_path, "static", "tile_templates")


# ------------------------------------------------
# Helpers
# ------------------------------------------------
def resolve_image_path(path: str | None) -> str | None:
    """Convert stored path or web path to real filesystem path."""
    if not path:
        return None

    # absolute path
    if os.path.isabs(path) and os.path.exists(path):
        return path

    stripped = path.lstrip("/")

    # relative to project root
    candidate = os.path.join(app.root_path, stripped)
    if os.path.exists(candidate):
        return candidate

    # fallback: relative to current working dir
    if os.path.exists(stripped):
        return os.path.abspath(stripped)

    return None


def get_poster_path() -> str | None:
    """
    Cover-page poster (first page).
    Prefer common names like cover.jpg / cover.png.
    """
    if not os.path.isdir(POSTER_DIR):
        return None

    preferred_names = [
        "cover.jpg", "cover.jpeg", "cover.png",
        "Cover.jpg", "Cover.JPG", "IMG_1306.JPG",
    ]
    for name in preferred_names:
        candidate = os.path.join(POSTER_DIR, name)
        if os.path.exists(candidate):
            print("Poster found (preferred):", candidate)
            return candidate

    # any jpg/png in posters
    for fname in os.listdir(POSTER_DIR):
        lower = fname.lower()
        if lower.endswith((".jpg", ".jpeg", ".png")):
            candidate = os.path.join(POSTER_DIR, fname)
            print("Poster found (fallback):", candidate)
            return candidate

    print("No poster image found in:", POSTER_DIR)
    return None


def get_tile_template_path() -> str | None:
    """
    Tile-page template (grey pamphlet with logo).
    Will pick ANY jpg/jpeg/png inside static/tile_templates.
    """
    if not os.path.isdir(TILE_TEMPLATE_DIR):
        print("Tile template dir does not exist:", TILE_TEMPLATE_DIR)
        return None

    for fname in os.listdir(TILE_TEMPLATE_DIR):
        lower = fname.lower()
        if lower.endswith((".jpg", ".jpeg", ".png")):
            candidate = os.path.join(TILE_TEMPLATE_DIR, fname)
            print("Tile template found:", candidate)
            return candidate

    print("No tile template image found in:", TILE_TEMPLATE_DIR)
    return None


# ------------------------------------------------
# Model
# NOTE: 'tags' column is now used as FINISH (e.g. GLOSSY, MATT)
# ------------------------------------------------
class Tile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200))
    sku = db.Column(db.String(100))      # kept in DB but not used in UI
    size = db.Column(db.String(100))
    price = db.Column(db.String(100))    # kept in DB but not used in UI
    description = db.Column(db.String(500))
    tags = db.Column(db.String(500))     # used as FINISH
    photo_path = db.Column(db.String(500))
    web_path = db.Column(db.String(500))


with app.app_context():
    db.create_all()


# ------------------------------------------------
# Home
# ------------------------------------------------
@app.route("/")
def home():
    tiles = Tile.query.order_by(Tile.id.desc()).all()
    return render_template("index.html", tiles=tiles)


# ------------------------------------------------
# Upload Tile – ONLY name, size, finish, description, photo
#   NAME and FINISH are stored in UPPERCASE
# ------------------------------------------------
@app.route("/upload_tile", methods=["GET", "POST"])
def upload_tile():
    if request.method == "POST":
        # Force uppercase for NAME and FINISH
        raw_name = request.form.get("name") or ""
        raw_finish = request.form.get("finish") or ""

        name = raw_name.upper()
        finish = raw_finish.upper()
        size = request.form.get("size") or ""
        description = request.form.get("description") or ""

        file = request.files.get("photo")
        photo_path = None
        web_path = None

        if file and file.filename:
            filename = f"{uuid.uuid4().hex}_{file.filename.replace(' ', '_')}"
            save_path = os.path.join(app.root_path, "static", "images", filename)
            file.save(save_path)

            photo_path = save_path
            web_path = "/static/images/" + filename

        tile = Tile(
            name=name,
            sku=None,                  # not used
            size=size,
            price=None,                # not used
            description=description,
            tags=finish,               # FINISH stored in tags
            photo_path=photo_path,
            web_path=web_path,
        )

        db.session.add(tile)
        db.session.commit()
        return redirect("/")

    return render_template("upload_tile.html")


# ------------------------------------------------
# Excel import  (also uppercases NAME and FINISH)
# ------------------------------------------------
def import_excel_with_images(filepath: str):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    image_map = {}
    for img in getattr(ws, "_images", []):
        try:
            anchor = img.anchor._from
            cell = f"{anchor.col_ref}{anchor.row}"
            image_map[cell] = img
        except Exception:
            continue

    for row in ws.iter_rows(min_row=2, values_only=False):
        name = row[0].value if row[0] else None
        sku = row[1].value if row[1] else None
        size = row[2].value if row[2] else None
        price = row[3].value if row[3] else None
        description = row[4].value if row[4] else None
        tags = row[5].value if row[5] else None   # FINISH here

        img_cell = row[6].coordinate if len(row) > 6 else None
        img_obj = image_map.get(img_cell)

        photo_path = None
        web_path = None

        if img_obj:
            try:
                img_bytes = img_obj._data()
                pil = Image.open(BytesIO(img_bytes))
                filename = f"{uuid.uuid4().hex}.png"
                save_path = os.path.join(app.root_path, "static", "images", filename)
                pil.save(save_path)

                photo_path = save_path
                web_path = "/static/images/" + filename
            except Exception as e:
                print("Image import error:", e)

        if name:
            # uppercase NAME and FINISH from Excel as well
            name_str = str(name).upper()
            finish_str = str(tags).upper() if tags is not None else None

            tile = Tile(
                name=name_str,
                sku=str(sku) if sku else None,
                size=str(size) if size else None,
                price=str(price) if price else None,
                description=str(description) if description else None,
                tags=finish_str,    # FINISH
                photo_path=photo_path,
                web_path=web_path
            )
            db.session.add(tile)

    db.session.commit()


@app.route("/upload_excel", methods=["GET", "POST"])
def upload_excel():
    if request.method == "GET":
        return render_template("upload_excel.html")

    file = request.files.get("excel_file")
    if not file:
        flash("No Excel file selected")
        return redirect("/upload_excel")

    filename = f"{uuid.uuid4().hex}_{file.filename}"
    save_path = os.path.join(app.root_path, "uploads", filename)
    file.save(save_path)
    import_excel_with_images(save_path)
    flash("Excel imported successfully")
    return redirect("/")


# ------------------------------------------------
# Delete tiles
# ------------------------------------------------
@app.route("/delete_tiles", methods=["POST"])
def delete_tiles():
    ids = request.form.getlist("tile_ids")
    if not ids:
        return redirect("/")

    for raw_id in ids:
        try:
            tid = int(raw_id)
        except ValueError:
            continue

        tile = Tile.query.get(tid)
        if not tile:
            continue

        if tile.photo_path:
            try:
                p = resolve_image_path(tile.photo_path)
                if p and os.path.exists(p):
                    os.remove(p)
            except Exception as e:
                print("Error removing image:", e)

        db.session.delete(tile)

    db.session.commit()
    return redirect("/")


# ------------------------------------------------
# PDF helpers
# ------------------------------------------------
def draw_cover_page(c: canvas.Canvas, client_name: str | None = None):
    """Cover page: full poster + client name on green band."""
    w, h = A4
    poster_path = get_poster_path()

    # full-page poster
    if poster_path:
        try:
            c.drawImage(
                ImageReader(poster_path),
                0,
                0,
                width=w,
                height=h,
                preserveAspectRatio=False,
                anchor="sw",
            )
        except Exception as e:
            print("Poster draw error:", e)

    # client name on green band
    if client_name:
        poster_height = h
        green_strip_y_start = poster_height * 0.08
        green_strip_y_end = poster_height * 0.18
        text_y = (green_strip_y_start + green_strip_y_end) / 2 - 24

        c.setFont("Helvetica-Bold", 28)
        c.setFillColorRGB(1, 1, 1)  # white
        c.drawCentredString(w / 2, text_y, f"Client: {client_name}")
        c.setFillColorRGB(0, 0, 0)


def draw_tile_detail_page(c: canvas.Canvas, tile: Tile):
    """
    Single tile page using grey template:
      - Template image as full background
      - ONE tile image centered
      - Size at top-right
      - DESIGN NAME :- {name}
      - FINISH :- {finish}  (finish from tags)
      All these texts use font size 18.
    """
    w, h = A4

    # background template
    tpl_path = get_tile_template_path()
    if tpl_path:
        try:
            c.drawImage(
                ImageReader(tpl_path),
                0,
                0,
                width=w,
                height=h,
                preserveAspectRatio=False,
                anchor="sw",
            )
        except Exception as e:
            print("Tile template draw error:", e)
    else:
        # fallback plain dark background
        c.setFillColorRGB(0.26, 0.28, 0.33)
        c.rect(0, 0, w, h, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)

    # --- IMPORTANT CHANGE HERE: prefer web_path over photo_path ---
    img_path = resolve_image_path(tile.web_path or tile.photo_path)
    print("DEBUG tile image:", tile.id, tile.web_path, tile.photo_path, "->", img_path)

    # tile image in center
    if img_path:
        try:
            max_w = w * 0.70
            max_h = h * 0.65
            x = (w - max_w) / 2
            y = (h - max_h) / 2
            c.drawImage(
                ImageReader(img_path),
                x,
                y,
                width=max_w,
                height=max_h,
                preserveAspectRatio=True,
                anchor="sw",
            )
        except Exception as e:
            print("Tile image draw error:", e)
    else:
        c.setFont("Helvetica-Bold", 18)
        c.setFillColorRGB(1, 1, 1)
        c.drawCentredString(w / 2, h / 2, "NO IMAGE FOUND")

    # text styling
    c.setFont("Helvetica-Bold", 18)
    c.setFillColorRGB(1, 1, 1)

    # size top-right
    size_text = (tile.size or "").upper()
    if size_text:
        c.drawRightString(w - 40, h - 40, size_text)

    # DESIGN NAME & FINISH bottom-left
    name_text = (tile.name or "").upper()
    finish_text = (tile.tags or "").upper()   # tags = FINISH

    y_base = 60
    if name_text:
        c.drawString(40, y_base, f"DESIGN NAME :- {name_text}")
        y_base -= 20
    if finish_text:
        c.drawString(40, y_base, f"FINISH :- {finish_text}")

    c.setFillColorRGB(0, 0, 0)


# ------------------------------------------------
# PDF: Single tile
# ------------------------------------------------
@app.route("/generate_pdf/<int:tile_id>")
def generate_pdf(tile_id):
    tile = Tile.query.get(tile_id)
    if not tile:
        return "Tile not found", 404

    client_name = request.args.get("client_name", "").strip() or None
    pdf_name = request.args.get("pdf_name", "").strip()

    # custom pdf file name
    if pdf_name:
        safe_name = pdf_name.replace(" ", "_")
        if not safe_name.lower().endswith(".pdf"):
            safe_name += ".pdf"
        pdf_path = safe_name
    else:
        pdf_path = f"tile_{tile_id}.pdf"

    c = canvas.Canvas(pdf_path, pagesize=A4)

    # cover
    draw_cover_page(c, client_name)
    c.showPage()

    # tile page
    draw_tile_detail_page(c, tile)

    c.save()
    return send_file(pdf_path, as_attachment=True)


# ------------------------------------------------
# PDF: Multiple tiles
# ------------------------------------------------
@app.route("/generate_pdf_multiple", methods=["POST"])
def generate_pdf_multiple():
    tile_ids = request.form.getlist("tile_ids")
    client_name = request.form.get("client_name", "").strip() or None
    pdf_name = request.form.get("pdf_name", "").strip()

    try:
        tile_ids = [int(i) for i in tile_ids]
    except Exception:
        return "Invalid tile ids", 400

    tiles = Tile.query.filter(Tile.id.in_(tile_ids)).order_by(Tile.id.desc()).all()
    if not tiles:
        return "No tiles selected", 400

    # custom pdf file name
    if pdf_name:
        safe_name = pdf_name.replace(" ", "_")
        if not safe_name.lower().endswith(".pdf"):
            safe_name += ".pdf"
        pdf_path = safe_name
    else:
        pdf_path = "tiles_selected.pdf"

    c = canvas.Canvas(pdf_path, pagesize=A4)

    # cover
    draw_cover_page(c, client_name)
    c.showPage()

    # tiles pages
    for idx, t in enumerate(tiles):
        draw_tile_detail_page(c, t)
        if idx < len(tiles) - 1:
            c.showPage()

    c.save()
    return send_file(pdf_path, as_attachment=True)


# ------------------------------------------------
# Run
# ------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)
